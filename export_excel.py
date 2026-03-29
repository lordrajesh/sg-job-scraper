"""
Excel Export — HK Job market intelligence
==========================================
- Dashboard: charts + key stats
- Cantonese detection (language requirements)
- Skills: fill blanks with top skills for the category
- All links clickable
"""

import re
import os
import sys
import argparse
import glob
from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from utils.analysis import (
    HARD_SKILLS, SOFT_SKILLS, MANDARIN_PATTERNS,
    extract_keywords, detect_mandarin, extract_max_salary, clean_salary,
)


# ==============================================================================
# STYLES
# ==============================================================================
BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
ORANGE = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
PURPLE = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
TEAL = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
STAT_BG = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BIG_NUM = Font(color="1F4E79", bold=True, size=26)
SMALL_LABEL = Font(color="888888", size=9)
SALARY_FONT = Font(bold=True, color="2E7D32")
EST_FONT = Font(italic=True, color="999999")
LINK_FONT = Font(color="1565C0", underline="single", size=10)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
PIE_COLORS = ["1F4E79", "2E7D32", "E65100", "4A148C", "00695C", "C62828", "1565C0", "FF6F00"]


def styled_header(ws, row, headers, fill):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = fill, WHITE_FONT, Alignment(horizontal="center"), BORDER


def add_link(ws, row, col, url):
    url = str(url).strip()
    if url and url.startswith("http"):
        cell = ws.cell(row=row, column=col)
        cell.value = '=HYPERLINK("' + url.replace('"', '%22') + '","Open")'
        cell.font = LINK_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    else:
        ws.cell(row=row, column=col, value="-").border = BORDER


# ==============================================================================
# MAIN
# ==============================================================================

def generate_excel(input_csv, output_path=None, merged_pairs=None, top_companies=None):
    df = pd.read_csv(input_csv)
    merged_pairs = merged_pairs or []
    top_companies = top_companies or []
    print(f"  Loaded {len(df)} listings")

    if output_path is None:
        output_path = os.path.join("output", f"hk_job_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    df["_max_sal"] = df["salary"].apply(extract_max_salary)
    df["_mandarin"] = df["description"].apply(detect_mandarin)
    df["_hard_skills"] = df["description"].apply(lambda d: extract_keywords(d, True))

    sal_valid = df[df["_max_sal"] > 0]["_max_sal"]
    all_median = int(sal_valid.median()) if len(sal_valid) else 25000

    all_hard = []
    all_soft = []
    for desc in df["description"].dropna():
        all_hard.extend(extract_keywords(desc, True))
        text = str(desc).lower()
        all_soft.extend(kw for kw in SOFT_SKILLS if re.search(rf'\b{re.escape(kw)}\b', text))
    hard_counts = Counter(all_hard).most_common(15)
    soft_counts = Counter(all_soft).most_common(5)

    top5_skills = ", ".join([kw for kw, _ in hard_counts[:5]]) if hard_counts else "excel, sql"

    df["_salary_display"] = df["salary"].apply(lambda s: clean_salary(s, all_median))
    df["_mandarin_display"] = df["_mandarin"].apply(lambda m: m if m else "Not required")
    df["_skills_display"] = df["_hard_skills"].apply(
        lambda s: ", ".join(s) if s else f"{top5_skills} (sector avg)"
    )
    df["_company_clean"] = df["company"].apply(
        lambda c: "Confidential" if not c or pd.isna(c) or c == "Unknown" else c
    )
    df["_type_clean"] = df["job_type"].fillna("-")

    mand_mask = df["_mandarin"] != ""
    mand_total = mand_mask.sum()

    wb = Workbook()

    # Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:K2")
    ws.cell(row=2, column=2, value="Hong Kong Job Market Dashboard").font = Font(color="1F4E79", bold=True, size=20)
    ws.merge_cells("B3:K3")
    ws.cell(row=3, column=2, value=f"{datetime.now().strftime('%d/%m/%Y')} | Sources: {', '.join(df['source'].unique())}").font = Font(color="888888", size=11)

    stats = [
        (str(len(df)), "Vacancies"),
        (str(df["_company_clean"].nunique()), "Companies"),
        (f"HK${all_median:,}", "Median Salary"),
        (str(mand_total), "Require Mandarin"),
    ]
    for i, (num, label) in enumerate(stats):
        col = 2 + i * 2
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        c1 = ws.cell(row=5, column=col, value=num)
        c1.font, c1.alignment, c1.fill = BIG_NUM, Alignment(horizontal="center"), STAT_BG
        ws.cell(row=5, column=col + 1).fill = STAT_BG
        c2 = ws.cell(row=6, column=col, value=label)
        c2.font, c2.alignment = SMALL_LABEL, Alignment(horizontal="center")

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file")
    args = parser.parse_args()
    if args.file:
        generate_excel(args.file)
    else:
        csvs = sorted(glob.glob("output/*.csv"), key=os.path.getmtime, reverse=True)
        if csvs:
            generate_excel(csvs[0])
        else:
            print("No CSV files")
